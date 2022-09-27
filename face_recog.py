from tkinter import *
from tkinter import font, ttk, messagebox, simpledialog
from PIL import ImageTk, Image 
import time
import customtkinter
import os
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

w=Tk()

#Using piece of code from old splash screen
width_of_window = 1000
height_of_window = 500
screen_width = w.winfo_screenwidth()
screen_height = w.winfo_screenheight()
x_coordinate = (screen_width/2)-(width_of_window/2)
y_coordinate = (screen_height/2)-(height_of_window/2)
w.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
w.configure(bg='#ED1B76')
w.overrideredirect(1) #for hiding titlebar



Frame(w, width=1000, height=500).place(x=0,y=0)
img= (Image.open("pic/1.png"))
resized_image_splash= img.resize((1000,500))
photo= ImageTk.PhotoImage(resized_image_splash)
# label = Label(w, image = photo)
# label.pack()

# label2=Label(w, text='Loading', fg='white', bg='#dab015') #decorate it 
# label2.configure(font=("Calibri", 17))
# label2.place(x=10,y=465)

#making animation

image_a=ImageTk.PhotoImage(Image.open('pic/dot2.png'))
image_b=ImageTk.PhotoImage(Image.open('pic/dot1.png'))


for i in range(5):
    # loading text
    label2=Label(w, text='Loading', fg='white', bg='#dab015') #decorate it 
    label2.configure(font=("Calibri", 17))
    label2.place(x=10,y=465)
    # background image
    label = Label(w, image = photo)
    label.pack()
    # loading dot animation
    l1=Label(w, image=image_a, border=0, relief=SUNKEN).place(x=90, y=480)
    l2=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=110, y=480)
    l3=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=130, y=480)
    l4=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=150, y=480)
    w.update_idletasks()
    time.sleep(0.5)

    # loading text
    label2=Label(w, text='Loading', fg='white', bg='#dab015') #decorate it 
    label2.configure(font=("Calibri", 17))
    label2.place(x=10,y=465)
    # background image
    label = Label(w, image = photo)
    label.pack()
    # loading dot animation
    l1=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=90, y=480)
    l2=Label(w, image=image_a, border=0, relief=SUNKEN).place(x=110, y=480)
    l3=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=130, y=480)
    l4=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=150, y=480)
    w.update_idletasks()
    time.sleep(0.5)

    # loading text
    label2=Label(w, text='Loading', fg='white', bg='#dab015') #decorate it 
    label2.configure(font=("Calibri", 17))
    label2.place(x=10,y=465)
    # background image
    label = Label(w, image = photo)
    label.pack()
    # loading dot animation
    l1=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=90, y=480)
    l2=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=110, y=480)
    l3=Label(w, image=image_a, border=0, relief=SUNKEN).place(x=130, y=480)
    l4=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=150, y=480)
    w.update_idletasks()
    time.sleep(0.5)

    # loading text
    label2=Label(w, text='Loading', fg='white', bg='#dab015') #decorate it 
    label2.configure(font=("Calibri", 17))
    label2.place(x=10,y=465)
    # background image
    label = Label(w, image = photo)
    label.pack()
    # loading dot animation
    l1=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=90, y=480)
    l2=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=110, y=480)
    l3=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=130, y=480)
    l4=Label(w, image=image_a, border=0, relief=SUNKEN).place(x=150, y=480)
    w.update_idletasks()
    time.sleep(0.5)

#new window to open
def new_win():

    customtkinter.set_appearance_mode("Light")  # Modes: "System" (standard), "Dark", "Light"
    customtkinter.set_default_color_theme("dark-blue")  # Themes: "blue" (standard), "green", "dark-blue"

    main_window=customtkinter.CTk()
    main_window.title('main window')
    # main_window.geometry('1000x500')
    main_window.state('zoomed')
    # main_window.rowconfigure(0, weight=2)
    # main_window.columnconfigure(1, weight=1)
    
    page1 = Frame(main_window)
    page2 = Frame(main_window)
    page3 = Frame(main_window)
    page4 = Frame(main_window)
    attendance_record = Frame(main_window)
    faculty_information = Frame(main_window)

    for frame in (page1, page2, page3, page4, attendance_record,faculty_information):
        frame.grid(row=0, column=0, sticky='nsew')

    def show_frame(frame):
        frame.tkraise()

    show_frame(page1)

    # ============= Page 1 Frame =========

        # open background image
    page1.pg1_image = Image.open('pic/2.png')
    page1.pg1_resize_image = page1.pg1_image.resize((1362, 692))
    page1.photo = ImageTk.PhotoImage(page1.pg1_resize_image)
    page1.pg1_bg_img_lb = Label(page1, image = page1.photo)
    page1.pg1_bg_img_lb.pack()
    
        # Login Button
    pg1_login_img_btn = PhotoImage(file = "pic/btn_login.png")
    pg1_button_admin = Button(page1,image=pg1_login_img_btn, borderwidth=0, bg='#1f2a76',command=lambda: show_frame(page2))
    pg1_button_admin.place(x=99, y=422)

        # Biomettric Button
    biometric_img_btn = PhotoImage(file = "pic/bttn_biometric.png")
    pg1_button_bio = Button(page1,image=biometric_img_btn, borderwidth=0,bg='#1f2a76', command=lambda: show_frame(page2))
    pg1_button_bio.place(x=99, y=256)

    # ============= Page 2 Frame =========

        # open background image
    page2.pg2_image = Image.open('pic/3.png')
    page2.pg2_resize_image = page2.pg2_image.resize((1362, 692))
    page2.photo = ImageTk.PhotoImage(page2.pg2_resize_image)
    page2.pg2_bg_img_lb = Label(page2, image = page2.photo)
    page2.pg2_bg_img_lb.pack()

        # Employee Button
    employee_img_btn = PhotoImage(file = "pic/btn_employee.png")
    pg2_button_bio = Button(page2,image=employee_img_btn, borderwidth=0,bg='#1f2a76', command=lambda: show_frame(page3))
    pg2_button_bio.place(x=99, y=256)

        # Admin Button
    admin_img_btn = PhotoImage(file = "pic/bttn_admin.png")
    pg2_button_admin = Button(page2,image=admin_img_btn, borderwidth=0, bg='#1f2a76',command=lambda: show_frame(page3))
    pg2_button_admin.place(x=99, y=422)

    # ======== Page 3 Admin Sign In Frame ===========

        # open background image
    page3.pg3_image = Image.open('pic/4.png')
    page3.pg3_resize_image = page3.pg3_image.resize((1362, 692))
    page3.photo = ImageTk.PhotoImage(page3.pg3_resize_image)
    page3.pg3_bg_img_lb = Label(page3, image = page3.photo)
    page3.pg3_bg_img_lb.pack()

        # Text Box Username and Password
    pg3_txtbox_username = Entry(page3, borderwidth=0, width=16, font=('Arial',30))
    pg3_txtbox_username.insert(0,"Username")
    pg3_txtbox_username.place(x=116, y=256, height=92)

    pg3_txtbox_pass = Entry(page3, borderwidth=0, width=16, font=('Arial', 30), show='*')
    pg3_txtbox_pass.insert(0,"Password")
    pg3_txtbox_pass.place(x=116, y=422, height=90)

        # Account verification
    def verify():
        try:
            with open("Account.txt", "r") as f:
                info = f.readlines()
                i  = 0
                for e in info:
                    u, p =e.split(",")
                    if u.strip() == pg3_txtbox_username.get() and p.strip() == pg3_txtbox_pass.get():
                        show_frame(page4)
                        i = 1
                        break
                if i==0:
                    messagebox.showinfo("Error", "Please provide correct username and password!!")
        except:
            messagebox.showinfo("Error", "Please provide correct username and password!!")

        # Login Button
    login_img_btn = PhotoImage(file = "pic/login.png")
    pg3_button = Button(page3, image=login_img_btn, borderwidth=0, bg='#1f2a76', command=verify)
    pg3_button.place(x=180, y=557)

        # show and hide Password
    def show_password():
        if  pg3_txtbox_pass.cget('show') =='*':
            pg3_txtbox_pass.configure(show='')
        else:
            pg3_txtbox_pass.configure(show='*')
    check_button = Checkbutton(page3, text="show password",bg="#1f2a76", command=show_password, font="Arial", activebackground="#1f2a76",)
    check_button.place(x=116,y=520)

    # ============= Page 4 Home  Frame =========

        # open background image
    page4.pg4_image = Image.open('pic/7.png')
    page4.pg4_resize_image = page4.pg4_image.resize((1362, 692))
    page4.photo = ImageTk.PhotoImage(page4.pg4_resize_image)
    page4.pg4_bg_img_lb = Label(page4, image = page4.photo)
    page4.pg4_bg_img_lb.pack()

        # Faculty Information Button
    faculty_info_btn = PhotoImage(file = "pic/faculty_info.png")
    pg4_button_faculty = customtkinter.CTkButton(master=page4,image=faculty_info_btn, text="" ,
                                                corner_radius=30, fg_color="#00436e",hover_color="#006699", command=lambda: show_frame(faculty_information))
    pg4_button_faculty.place(x=100, y=314, height=134,width=562)

        # Attendace Record Button
    att_rec_btn = PhotoImage(file = "pic/attendance_rec.png")
    pg4_button_att_rec = customtkinter.CTkButton(master=page4,image=att_rec_btn, text="" ,
                                                corner_radius=30, fg_color="#cc9900",hover_color="#fdca34", command=lambda: show_frame(attendance_record))
    pg4_button_att_rec.place(x=100, y=469, height=178,width=330)

        # About Button
    about_btn = PhotoImage(file = "pic/about.png")
    pg4_button_photo = customtkinter.CTkButton(master=page4,image=about_btn, text="" ,
                                                corner_radius=30, fg_color="#cc9900",hover_color="#fdca34", command=lambda: show_frame(page2))
    pg4_button_photo.place(x=456, y=469, height=178,width=197)

        # Photo Button
    photo_btn = PhotoImage(file = "pic/photo.png")
    pg4_button_train_img = customtkinter.CTkButton(master=page4,image=photo_btn, text="" ,
                                                corner_radius=30, fg_color="#cc9900",hover_color="#fdca34", command=lambda: show_frame(page2))
    pg4_button_train_img.place(x=683, y=315, height=333,width=348)

        # Developers Button
    developers_btn = PhotoImage(file = "pic/developers.png")
    pg4_button_developers = customtkinter.CTkButton(master=page4,image=developers_btn, text="" ,
                                                corner_radius=20, fg_color="#00436e",hover_color="#006699", command=lambda: show_frame(page2))
    pg4_button_developers.place(x=1048, y=314, height=134,width=246)

        # Logout Button
    logout_btn = PhotoImage(file = "pic/logout.png")
    pg4_button_logout = customtkinter.CTkButton(master=page4,image=logout_btn, text="" ,
                                                corner_radius=30, fg_color="#f0f0f0",hover_color="#6699cc", command=lambda: show_frame(page2))
    pg4_button_logout.place(x=1075, y=469, height=178,width=197)

    # ============= Attendace Record Frame =========

        # open background image
    attendance_record.att_rec_image = Image.open('pic/8.png')
    attendance_record.att_rec_resize_image = attendance_record.att_rec_image.resize((1362, 692))
    attendance_record.photo = ImageTk.PhotoImage(attendance_record.att_rec_resize_image)
    attendance_record.att_rec_bg_img_lb = Label(attendance_record, image = attendance_record.photo)
    attendance_record.att_rec_bg_img_lb.pack()

        # Mathematics Faculty Button
    math_fac_btn = PhotoImage(file = "pic/math_faculty.png")
    att_rec_button_math_fac = customtkinter.CTkButton(master=attendance_record,image=math_fac_btn, text="" ,
                                                corner_radius=20,bg='#ffffff', fg_color="#00436e",hover_color="#006699", command=lambda: show_frame(page2))
    att_rec_button_math_fac.place(x=254, y=254, height=99,width=413)

        # Psychology Faculty Button
    psyc_fac_btn = PhotoImage(file = "pic/psyc_faculty.png")
    att_rec_button_psyc_fac = customtkinter.CTkButton(master=attendance_record,image=psyc_fac_btn, text="" ,
                                                corner_radius=20,bg='#ffffff', fg_color="#00436e",hover_color="#006699", command=lambda: show_frame(page2))
    att_rec_button_psyc_fac.place(x=731, y=254, height=99,width=413)

        # ITE Faculty Button
    ite_fac_btn = PhotoImage(file = "pic/ite_faculty.png")
    att_rec_button_ite_fac = customtkinter.CTkButton(master=attendance_record,image=ite_fac_btn, text="" ,
                                                corner_radius=20,bg='#ffffff', fg_color="#00436e",hover_color="#006699", command=lambda: show_frame(page2))
    att_rec_button_ite_fac.place(x=254, y=394, height=99,width=413)

        # Applied Psychology Faculty Button
    app_psyc_fac_btn = PhotoImage(file = "pic/psyc_faculty.png")
    att_rec_button_app_psyc_fac = customtkinter.CTkButton(master=attendance_record,image=app_psyc_fac_btn, text="" ,
                                                corner_radius=20,bg='#ffffff', fg_color="#00436e",hover_color="#006699", command=lambda: show_frame(page2))
    att_rec_button_app_psyc_fac.place(x=731, y=394, height=99,width=413)

    # ============= Faculty Information Frame =========

        # open background image
    faculty_information.fac_inf_image = Image.open('pic/10.png')
    faculty_information.fac_inf_resize_image = faculty_information.fac_inf_image.resize((1362, 692))
    faculty_information.photo = ImageTk.PhotoImage(faculty_information.fac_inf_resize_image)
    faculty_information.fac_inf_bg_img_lb = Label(faculty_information, image = faculty_information.photo)
    faculty_information.fac_inf_bg_img_lb.pack()

        # ComboBox College Department
    department_combobox = ttk.Combobox(faculty_information, values=["Mathematics", "ITE", "Psychology", "Applied Physics"])
    department_combobox.place(x=382, y=202, width=200)
    
        # Entry Employee Number
    employee_num_fac_inf = Entry(faculty_information)
    employee_num_fac_inf.place(x=319, y=284, width=125)

        # Entry Employee Name
    employee_name_fac_inf = Entry(faculty_information)
    employee_name_fac_inf.place(x=560, y=284, width=125)

        # ComboBox Gender
    gender_combobox_fac_inf = ttk.Combobox(faculty_information, values=["Male", "Female"])
    gender_combobox_fac_inf.place(x=319, y=316, width=125)

        # Entry Age
    age_fac_inf = Entry(faculty_information)
    age_fac_inf.place(x=560, y=316, width=125)

        # Entry E-mail
    email_fac_inf = Entry(faculty_information)
    email_fac_inf.place(x=319, y=348, width=125)

        # Entry Contact Number
    con_num_fac_inf = Entry(faculty_information)
    con_num_fac_inf.place(x=560, y=348, width=125)

        # Entry Address
    address_fac_inf = Entry(faculty_information)
    address_fac_inf.place(x=319, y=380, width=125)

        # Search Entry
    def filterTreeView(*args):
        ItemOnTreeView = data_table.get_children()

        search = search_ent_val.get()

        for eachItem in ItemOnTreeView:
            if search in data_table.item(eachItem)['values'][2]:
                search_val = data_table.item(eachItem)['values']
                data_table.delete(eachItem)

                data_table.insert("", 1, values=search_val)

    search_ent_val = StringVar()

    search_fac_inf = Entry(faculty_information, textvariable = search_ent_val)
    search_fac_inf.place(x=850, y=202, width=200)

    search_ent_val.trace("w", filterTreeView)

        # open and connect the excel to tha data table
    openfile = openpyxl.load_workbook("data/faculty_data.xlsx")
    select_sheet1 = openfile.active

    list_col_header = select_sheet1.iter_rows(min_row=1,max_row=1,values_only=True)
    list_data_set = select_sheet1.iter_rows(min_row=2,values_only=True)
    list_col_header = [r for r in list_col_header]
    list_data_set = [r for r in list_data_set]
    openfile.close()

        # Data Table "TreeView"
    scrollbarx = Scrollbar(faculty_information, orient=HORIZONTAL)
    scrollbarx.place(x=730, y=584, width=465)
    scrollbary = Scrollbar(faculty_information, orient=VERTICAL)
    scrollbary.place(x=1180, y=284, height=300)

    style = ttk.Style()
    style.configure("Treeview.Heading", font=("yu gothic ui", 10, "bold"))

    data_table = ttk.Treeview(faculty_information, selectmode = 'browse')
    data_table.place(x=730, y=284, width=450, height=300)
    data_table.configure(yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)

    scrollbarx.configure(command=data_table.xview)
    scrollbary.configure(command=data_table.yview)

    data_table['height'] = 8
    data_table['show'] = 'headings'
    data_table["columns"] = list_col_header[0]

    # Display the Heading 
    for i in list_col_header[0]:
        data_table.column(i, width = 100, anchor = 'c')
    for i in list_col_header[0]:
        data_table.heading(i, text = i)

    # Display the list Data
    for data in list_data_set:
        data_table.insert("", 'end', iid = data[0], values = data)

        # Clear Text Field
    def clear():
        department_combobox.delete(0, END)
        employee_num_fac_inf.delete(0, END)
        gender_combobox_fac_inf.delete(0, END)
        email_fac_inf.delete(0, END)
        address_fac_inf.delete(0, END)
        employee_name_fac_inf.delete(0, END)
        age_fac_inf.delete(0, END)
        con_num_fac_inf.delete(0, END)

        # Add Faculty Button
    def Save_Data():
        save_college_department = department_combobox.get()
        save_employee_number = employee_num_fac_inf.get()
        save_gender = gender_combobox_fac_inf.get()
        save_email = email_fac_inf.get()
        save_address = address_fac_inf.get()
        save_employee_name = employee_name_fac_inf.get()
        save_age = age_fac_inf.get()
        save_contact_number = con_num_fac_inf.get()

        if save_college_department and save_employee_number and save_gender and save_email and save_address and save_employee_name and save_age and save_contact_number:

            # print("College Department:", save_college_department)
            # print("Employee No.: ", save_employee_number, "Employee Name: ", save_employee_name)
            # print("Gender: ", save_gender, "Age: ", save_age)
            # print("Email: ", save_email, "Contact Number: ", save_contact_number)
            # print("Address:", save_address)
            # print("------------------------------------------")

            filepath = pathlib.Path("data/faculty_data.xlsx")
            if filepath.exists():
                pass 
            else:
                filepath = Workbook()
                sheet = filepath.active
                sheet["A1"] = "Employee No."
                sheet["B1"] = "Email"
                sheet["C1"] = "Employee Name"
                sheet["D1"] = "Gender"
                sheet["E1"] = "Age"
                sheet["F1"] = "Contact Number"
                sheet["G1"] = "Address"
                sheet["H1"] = "College Department"

                filepath.save("data/faculty_data.xlsx")

            file = openpyxl.load_workbook("data/faculty_data.xlsx")
            sheet1 = file.active

            Found = False

            for i in range(2,(sheet1.max_row)+1):
                if((save_employee_number==sheet1['A'+str(i)].value) and (save_email==sheet1['B'+str(i)].value)):
                    Found = True
                    break
                else:
                    Found = False
                
            if(Found==True):
                messagebox.showinfo("Error", "Employee Number or Email Exists!!")
            else:
                messagebox.showinfo("Success", "Data Added!")
                lastRow = str((sheet1.max_row)+1)
                sheet1['A'+lastRow] = save_employee_number
                sheet1['B'+lastRow] = save_email
                sheet1['C'+lastRow] = save_employee_name
                sheet1['D'+lastRow] = save_gender
                sheet1['E'+lastRow] = save_age
                sheet1['F'+lastRow] = save_contact_number
                sheet1['G'+lastRow] = save_address
                sheet1['H'+lastRow] = save_college_department

            file.save("data/faculty_data.xlsx")
            clear()

        else:
            messagebox.showwarning(title="Error", message="Fill up all the data are required.")

    add_fac_btn = PhotoImage(file = "pic/btn_add_faculty.png")
    add_button_fac_inf = customtkinter.CTkButton(master=faculty_information,image=add_fac_btn, text="" ,
                                                corner_radius=6,bg='#ffffff', fg_color="#00436e",hover_color="#006699", command= Save_Data)
    add_button_fac_inf.place(x=380, y=506, height=32,width=131)

        # Update Button
    def Update():
        print("update")
    update_btn = PhotoImage(file = "pic/btn_update.png")
    add_button_update = customtkinter.CTkButton(master=faculty_information,image=update_btn, text="" ,
                                                corner_radius=6,bg='#ffffff', fg_color="#00436e",hover_color="#006699", command= Update)
    add_button_update.place(x=212, y=506, height=32,width=131)

        # Reset Button
    reset_btn = PhotoImage(file = "pic/btn_reset.png")
    add_button_reset = customtkinter.CTkButton(master=faculty_information,image=reset_btn, text="" ,
                                                corner_radius=6,bg='#ffffff', fg_color="#00436e",hover_color="#006699", command= clear)
    add_button_reset.place(x=548, y=506, height=32,width=131)

    main_window.mainloop()

w.destroy()
new_win()
w.mainloop()